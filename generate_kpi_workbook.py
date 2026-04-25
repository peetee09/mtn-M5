#!/usr/bin/env python3
"""
generate_kpi_workbook.py
========================
Python / openpyxl alternative to the VBA macro.
Generates the complete Warehouse / Dispatch KPI workbook.

Requirements:
    pip install openpyxl

Usage:
    python generate_kpi_workbook.py
    # -> creates  KPI_Workbook.xlsx  in the current directory

Compatibility: Excel 2016 and later.
Unprotect password: KPI2024
"""

from __future__ import annotations

import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference

# ---------------------------------------------------------------------------
# Colour helpers
# ---------------------------------------------------------------------------
GREEN_HEX   = "FF50C878"
AMBER_HEX   = "FFFFA500"
RED_HEX     = "FFFF0000"
HEADER_HEX  = "FF191970"   # midnight blue
WHITE_HEX   = "FFFFFFFF"
LGREY_HEX   = "FFF2F2F2"
BLUE_HEX    = "FF0070C0"

# Dashboard / Excel-style palette
PAGE_BG_HEX  = "FFF2F2F2"   # very light grey page background
CARD_BG_HEX  = "FFFFFFFF"   # white card value area
SECN_HEX     = "FF2E4057"   # dark slate for section header bars
NAV_HEX      = "FF595959"   # dark grey for navigation bar
STEEL_HEX    = "FF4472C4"   # standard Excel blue for RAG legend bar
TEXT_DK_HEX  = "FF333333"   # near-black body text
TEXT_MD_HEX  = "FF666666"   # medium grey labels
LGREY2_HEX   = "FFD9D9D9"   # thin border grey

UNPROTECT_PW = "KPI2024"
# NOTE: This password is intentionally visible in the source code.
# It is a shared workbook protection password documented for all operators
# (not a user authentication credential). Change it in CONFIG if required.


def _hex_font_color(hex_color: str) -> str:
    """Strip the two-character alpha prefix from an ARGB hex string for Font.color."""
    return hex_color[2:]


def fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def bold_font(color: str = "FF000000", size: int = 11) -> Font:
    return Font(bold=True, color=color, size=size)


def header_style(ws, row: int, cols: List[str], start_col: int = 1) -> None:
    """Write and style a header row."""
    for c_idx, name in enumerate(cols, start=start_col):
        cell = ws.cell(row=row, column=c_idx, value=name)
        cell.font = bold_font(WHITE_HEX)
        cell.fill = fill(HEADER_HEX)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def add_table(ws, ref: str, name: str, style: str = "TableStyleMedium2",
              show_filter: bool = True) -> Table:
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    return tbl


def col_letter(n: int) -> str:
    return get_column_letter(n)


def add_shift_validation(ws, col_letter_str: str, first_row: int = 2) -> None:
    dv = DataValidation(
        type="list", formula1='"Day,Night"',
        allow_blank=True, showDropDown=False,
        showInputMessage=True, promptTitle="Shift",
        prompt="Select Day or Night",
        showErrorMessage=True, errorTitle="Invalid",
        error="Please select Day or Night"
    )
    dv.sqref = f"{col_letter_str}{first_row}:{col_letter_str}10000"
    ws.add_data_validation(dv)


def freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_config(wb: Workbook) -> None:
    ws = wb.create_sheet("CONFIG")
    ws.sheet_properties.tabColor = "808080"

    ws["A1"] = "CONFIG - DO NOT DELETE OR RENAME THIS SHEET"
    ws["A1"].font = Font(bold=True, color=_hex_font_color(RED_HEX), size=12)

    # Shift table
    ws["A3"] = "SHIFTS"
    ws["A3"].font = Font(bold=True)
    shift_hdrs = ["ShiftName", "StartTime", "EndTime", "DaysPerWeek"]
    header_style(ws, 4, shift_hdrs)

    shift_data = [
        ("Day",   "08:00", "16:00", 6),
        ("Night", "16:00", "00:00", 6),
    ]
    for r_off, row in enumerate(shift_data, start=5):
        for c_off, val in enumerate(row, start=1):
            ws.cell(row=r_off, column=c_off, value=val)

    add_table(ws, "A4:D6", "tblConfig_Shifts", "TableStyleMedium2")

    # KPI rules
    ws["A9"] = "KPI RULES"
    ws["A9"].font = Font(bold=True)
    rule_hdrs = ["RuleName", "Value", "Description"]
    header_style(ws, 10, rule_hdrs)

    rules = [
        ("HRP_MaxDaysToShow",  1,    "Max days since available for city"),
        ("Packed_MaxAgeDays",  2,    "Max age (days) before packed item is overdue"),
        ("Audit_SampleSize",   20,   "Required number of audits per shift"),
        ("Amber_Threshold",    0.9,  "Performance % for Amber RAG"),
        ("Green_Threshold",    1,    "Performance % for Green RAG"),
    ]
    for r_off, row in enumerate(rules, start=11):
        for c_off, val in enumerate(row, start=1):
            ws.cell(row=r_off, column=c_off, value=val)

    add_table(ws, "A10:C15", "tblConfig_Rules", "TableStyleMedium2")

    # Area list
    ws["E3"] = "AREA LIST"
    ws["E3"].font = Font(bold=True)
    header_style(ws, 4, ["AreaName"], start_col=5)
    areas = ["Auditing", "Manual handover", "Dispatch sealing"]
    for r_off, area in enumerate(areas, start=5):
        ws.cell(row=r_off, column=5, value=area)

    add_table(ws, "E4:E7", "tblConfig_Areas", "TableStyleMedium2")

    _autofit(ws)


def build_in_packed(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_PACKED")
    ws.sheet_properties.tabColor = "0070C0"

    src_cols = [
        "LPN", "PALLET", "STORE", "DIVISION", "FACILITY_NAME",
        "STORE_STATUS", "PICK_LOC", "UNITS", "LAST_UPDATE",
        "DAYS_SINCE_CLOSE", "LOCKS", "TOT_COST", "TOT_RETAIL", "LAST_PACKED"
    ]
    calc_cols = ["AgeDays", "IsShipped", "PackedStatus", "ActionFlag"]
    all_cols = src_cols + calc_cols

    header_style(ws, 1, all_cols)

    # Sample data row
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    sample = [
        "LPN001", "PLT001", "STORE01", "DIV01", "FACILITY A",
        "OPEN", "LOC001", 10, datetime.datetime.now(),
        1, 0, 100.0, 150.0, datetime.datetime.combine(yesterday, datetime.time(10, 0))
    ]
    n_src = len(src_cols)
    for c_off, val in enumerate(sample, start=1):
        ws.cell(row=2, column=c_off, value=val)

    # Calculated columns (stored as formulas in row 2; table will propagate)
    calc_start = n_src + 1   # column 15
    ws.cell(row=2, column=calc_start).value     = "=TODAY()-INT([@LAST_PACKED])"
    ws.cell(row=2, column=calc_start + 1).value = "=COUNTIF(tblShipped[LPN],[@LPN])>0"
    ws.cell(row=2, column=calc_start + 2).value = \
        '=IF([@IsShipped],"Shipped",IF([@AgeDays]>2,"Overdue","Pending"))'
    ws.cell(row=2, column=calc_start + 3).value = "=AND(NOT([@IsShipped]),[@AgeDays]>2)"

    last_col = col_letter(len(all_cols))
    add_table(ws, f"A1:{last_col}2", "tblPacked", "TableStyleMedium6")

    # Conditional formatting: PackedStatus column (col calc_start+2)
    ps_col = col_letter(calc_start + 2)
    ws.conditional_formatting.add(
        f"{ps_col}2:{ps_col}10000",
        Rule(type="containsText", operator="containsText", text="Overdue",
             dxf=DifferentialStyle(fill=fill(RED_HEX), font=Font(color=_hex_font_color(WHITE_HEX))))
    )
    ws.conditional_formatting.add(
        f"{ps_col}2:{ps_col}10000",
        Rule(type="containsText", operator="containsText", text="Shipped",
             dxf=DifferentialStyle(fill=fill(GREEN_HEX)))
    )
    ws.conditional_formatting.add(
        f"{ps_col}2:{ps_col}10000",
        Rule(type="containsText", operator="containsText", text="Pending",
             dxf=DifferentialStyle(fill=fill(AMBER_HEX)))
    )

    freeze(ws)
    _autofit(ws)


def build_in_hrp(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_HRP")
    ws.sheet_properties.tabColor = "00B0F0"

    src_cols = [
        "CATEGORY", "DAYS_SINCE_AVAILABLE_FOR_CITY", "DATE_SCANNED_TO_CITY",
        "OLPN", "XREF_OLPN", "STORE_NUMBER", "STORE_NAME",
        "CARTON_LOCKS", "LPN_STATUS", "STORE_ACK_STATUS",
        "UNITS", "COST_VALUE", "RETAIL_VALUE"
    ]
    calc_cols = ["CartonID", "IsAcknowledged", "IncludeInHRP", "AgeBucket"]
    all_cols = src_cols + calc_cols

    header_style(ws, 1, all_cols)

    today = datetime.date.today()
    sample_src = [
        "HRP", 0, datetime.datetime.now(),
        "OLPN001", "", "STORE01", "Store One",
        0, "AVAILABLE", "N", 5, 50.0, 75.0
    ]
    for c_off, val in enumerate(sample_src, start=1):
        ws.cell(row=2, column=c_off, value=val)

    calc_start = len(src_cols) + 1
    ws.cell(row=2, column=calc_start).value     = '=IF([@OLPN]<>"",[@OLPN],[@XREF_OLPN])'
    ws.cell(row=2, column=calc_start + 1).value = '=[@STORE_ACK_STATUS]="Y"'
    ws.cell(row=2, column=calc_start + 2).value = \
        '=AND([@STORE_ACK_STATUS]="N",[@DAYS_SINCE_AVAILABLE_FOR_CITY]<=1)'
    ws.cell(row=2, column=calc_start + 3).value = \
        '=IF([@DAYS_SINCE_AVAILABLE_FOR_CITY]=0,"<24h","24h+")'

    last_col = col_letter(len(all_cols))
    add_table(ws, f"A1:{last_col}2", "tblHRP", "TableStyleMedium7")

    # Data validation on STORE_ACK_STATUS
    ack_col = col_letter(src_cols.index("STORE_ACK_STATUS") + 1)
    dv_ack = DataValidation(
        type="list", formula1='"Y,N"', allow_blank=True,
        showDropDown=False, showInputMessage=True,
        promptTitle="Ack Status", prompt="Enter Y or N"
    )
    dv_ack.sqref = f"{ack_col}2:{ack_col}10000"
    ws.add_data_validation(dv_ack)

    # CF: IncludeInHRP = TRUE  (column calc_start+2)
    # formula=["TRUE"] generates <formula>TRUE</formula> in XML, which Excel
    # evaluates as the boolean constant TRUE (not the text string "TRUE").
    inc_col = col_letter(calc_start + 2)
    ws.conditional_formatting.add(
        f"{inc_col}2:{inc_col}10000",
        Rule(type="cellIs", operator="equal", formula=["TRUE"],
             dxf=DifferentialStyle(fill=fill(RED_HEX), font=Font(color=_hex_font_color(WHITE_HEX))))
    )

    freeze(ws)
    _autofit(ws)


def build_in_shipped(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_SHIPPED_LPNS")
    ws.sheet_properties.tabColor = "00B050"

    cols = ["BusinessDate", "ShiftName", "LPN", "EnteredBy", "Notes", "DupFlag"]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2, value="Day")
    ws.cell(row=2, column=3, value="LPN001")
    ws.cell(row=2, column=4, value="User1")
    ws.cell(row=2, column=5, value="")
    ws.cell(row=2, column=6).value = \
        "=COUNTIFS([BusinessDate],[@BusinessDate],[ShiftName],[@ShiftName],[LPN],[@LPN])>1"

    add_table(ws, "A1:F2", "tblShipped", "TableStyleMedium9")

    # CF: DupFlag = TRUE
    ws.conditional_formatting.add(
        "F2:F10000",
        Rule(type="cellIs", operator="equal", formula=["TRUE"],
             dxf=DifferentialStyle(fill=fill(AMBER_HEX)))
    )

    add_shift_validation(ws, "B")
    freeze(ws)
    _autofit(ws)


def build_in_staffing(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_STAFFING")
    ws.sheet_properties.tabColor = "00B050"

    cols = ["BusinessDate", "ShiftName", "Area", "StaffAvailable"]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2, value="Day")
    ws.cell(row=2, column=3, value="Dispatch sealing")
    ws.cell(row=2, column=4, value=5)

    add_table(ws, "A1:D2", "tblStaffing", "TableStyleMedium4")

    add_shift_validation(ws, "B")

    dv_area = DataValidation(
        type="list",
        formula1='"Auditing,Manual handover,Dispatch sealing"',
        allow_blank=True, showDropDown=False
    )
    dv_area.sqref = "C2:C10000"
    ws.add_data_validation(dv_area)

    freeze(ws)
    _autofit(ws)


def build_in_targets(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_TARGETS_DAILY")
    ws.sheet_properties.tabColor = "00B050"

    cols = ["BusinessDate", "ShiftName", "TargetPerPersonPerShift"]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2, value="Day")
    ws.cell(row=2, column=3, value=50)

    add_table(ws, "A1:C2", "tblTargetsDaily", "TableStyleMedium4")
    add_shift_validation(ws, "B")
    freeze(ws)
    _autofit(ws)


def build_in_audit_log(wb: Workbook) -> None:
    ws = wb.create_sheet("IN_AUDIT_LOG")
    ws.sheet_properties.tabColor = "70AD47"

    cols = ["BusinessDate", "ShiftName", "Area", "AuditCount", "Notes"]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2, value="Day")
    ws.cell(row=2, column=3, value="Auditing")
    ws.cell(row=2, column=4, value=0)
    ws.cell(row=2, column=5, value="")

    add_table(ws, "A1:E2", "tblAuditLog", "TableStyleMedium4")
    add_shift_validation(ws, "B")

    dv_area = DataValidation(
        type="list",
        # Cross-sheet structured references for data validation are the unreliable
        # form here. The stable explicit reference is an absolute range such as
        # =CONFIG!$E$5:$E$7 (used by the VBA implementation), but this openpyxl
        # path keeps the area list inline for compatibility. Keep these values in
        # sync with tblConfig_Areas in build_config().
        formula1='"Auditing,Manual handover,Dispatch sealing"',
        allow_blank=True, showDropDown=False
    )
    dv_area.sqref = "C2:C10000"
    ws.add_data_validation(dv_area)

    freeze(ws)
    _autofit(ws)


def build_t_dispatch_kpi(wb: Workbook) -> None:
    ws = wb.create_sheet("T_DISPATCH_KPI")
    ws.sheet_properties.tabColor = "FFC000"

    cols = [
        "BusinessDate", "ShiftName", "ShippedCartons", "TotalStaff",
        "TargetPerPerson", "ExpectedCartons", "PerformancePct", "RAG",
        "AuditCount", "AuditTarget", "AuditPct", "HRPOpen", "PackedOverdue"
    ]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2, value="Day")
    ws.cell(row=2, column=3).value = \
        "=COUNTIFS(tblShipped[BusinessDate],[@BusinessDate],tblShipped[ShiftName],[@ShiftName])"
    ws.cell(row=2, column=4).value = \
        "=SUMIFS(tblStaffing[StaffAvailable],tblStaffing[BusinessDate],[@BusinessDate],tblStaffing[ShiftName],[@ShiftName])"
    # Array formula (INDEX/MATCH across two criteria)
    ws.cell(row=2, column=5).value = \
        ("=IFERROR(INDEX(tblTargetsDaily[TargetPerPersonPerShift],"
         "MATCH(1,(tblTargetsDaily[BusinessDate]=[@BusinessDate])"
         "*(tblTargetsDaily[ShiftName]=[@ShiftName]),0)),0)")
    ws.cell(row=2, column=6).value = "=[@TotalStaff]*[@TargetPerPerson]"
    ws.cell(row=2, column=7).value = "=IFERROR([@ShippedCartons]/[@ExpectedCartons],0)"
    ws.cell(row=2, column=8).value = \
        '=IF([@PerformancePct]>=1,"Green",IF([@PerformancePct]>=0.9,"Amber","Red"))'
    # Audit metrics
    ws.cell(row=2, column=9).value = \
        "=SUMIFS(tblAuditLog[AuditCount],tblAuditLog[BusinessDate],[@BusinessDate],tblAuditLog[ShiftName],[@ShiftName])"
    ws.cell(row=2, column=10).value = \
        '=IFERROR(INDEX(tblConfig_Rules[Value],MATCH("Audit_SampleSize",tblConfig_Rules[RuleName],0)),20)'
    ws.cell(row=2, column=11).value = "=IFERROR([@AuditCount]/[@AuditTarget],0)"
    # Live snapshot counts (reflect current state of source tables)
    ws.cell(row=2, column=12).value = "=COUNTIF(tblHRP[IncludeInHRP],TRUE)"
    ws.cell(row=2, column=13).value = "=COUNTIF(tblPacked[ActionFlag],TRUE)"

    ws.cell(row=2, column=7).number_format = "0.0%"
    ws.cell(row=2, column=11).number_format = "0.0%"

    add_table(ws, "A1:M2", "tblDispatchKPI", "TableStyleMedium2")

    # RAG conditional formatting
    rag_col = "H"
    for text, hex_color in [("Green", GREEN_HEX), ("Amber", AMBER_HEX), ("Red", RED_HEX)]:
        ws.conditional_formatting.add(
            f"{rag_col}2:{rag_col}10000",
            Rule(type="containsText", operator="containsText", text=text,
                 dxf=DifferentialStyle(fill=fill(hex_color)))
        )

    # AuditPct data bar (column K = 11)
    from openpyxl.formatting.rule import DataBarRule
    ws.conditional_formatting.add(
        "K2:K10000",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=1,
                    color="FF70AD47")
    )

    freeze(ws)
    _autofit(ws)


def build_t_dispatch_daily(wb: Workbook) -> None:
    ws = wb.create_sheet("T_DISPATCH_DAILY")
    ws.sheet_properties.tabColor = "FFC000"

    cols = ["BusinessDate", "TotalShipped", "TotalStaff", "TotalExpected", "DailyPerformancePct"]
    header_style(ws, 1, cols)

    ws.cell(row=2, column=1, value=datetime.date.today())
    ws.cell(row=2, column=2).value = \
        "=SUMIF(tblDispatchKPI[BusinessDate],[@BusinessDate],tblDispatchKPI[ShippedCartons])"
    ws.cell(row=2, column=3).value = \
        "=SUMIF(tblDispatchKPI[BusinessDate],[@BusinessDate],tblDispatchKPI[TotalStaff])"
    ws.cell(row=2, column=4).value = \
        "=SUMIF(tblDispatchKPI[BusinessDate],[@BusinessDate],tblDispatchKPI[ExpectedCartons])"
    ws.cell(row=2, column=5).value = "=IFERROR([@TotalShipped]/[@TotalExpected],0)"
    ws.cell(row=2, column=5).number_format = "0.0%"

    add_table(ws, "A1:E2", "tblDispatchDaily", "TableStyleMedium2")

    freeze(ws)
    _autofit(ws)


def build_action_hrp(wb: Workbook) -> None:
    ws = wb.create_sheet("ACTION_HRP")
    ws.sheet_properties.tabColor = "C00000"

    cols = [
        "CATEGORY", "CartonID", "STORE_NUMBER", "STORE_NAME",
        "DAYS_SINCE_AVAILABLE_FOR_CITY", "STORE_ACK_STATUS",
        "UNITS", "COST_VALUE",
        "Owner", "ContactedCity", "ContactTime", "NextStep"
    ]
    for c_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = bold_font(WHITE_HEX)
        cell.fill = fill("FFC00000")
        cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="-- Populate via 'POPULATE ACTION SHEETS' button on DASHBOARD --")
    ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

    # Validation on ContactedCity
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.sqref = "J2:J10000"
    ws.add_data_validation(dv)

    freeze(ws)
    _autofit(ws)


def build_action_packed(wb: Workbook) -> None:
    ws = wb.create_sheet("ACTION_PACKED")
    ws.sheet_properties.tabColor = "C00000"

    cols = [
        "LPN", "STORE", "DIVISION", "AgeDays",
        "PackedStatus", "UNITS",
        "HoldReason", "Owner", "PlannedShipTime"
    ]
    for c_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = bold_font(WHITE_HEX)
        cell.fill = fill("FFC00000")
        cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="-- Populate via 'POPULATE ACTION SHEETS' button on DASHBOARD --")
    ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

    # CF on PackedStatus column (col E)
    ws.conditional_formatting.add(
        "E2:E10000",
        Rule(type="containsText", operator="containsText", text="Overdue",
             dxf=DifferentialStyle(fill=fill(RED_HEX), font=Font(color=_hex_font_color(WHITE_HEX))))
    )
    ws.conditional_formatting.add(
        "E2:E10000",
        Rule(type="containsText", operator="containsText", text="Pending",
             dxf=DifferentialStyle(fill=fill(AMBER_HEX)))
    )

    freeze(ws)
    _autofit(ws)


def build_history(wb: Workbook) -> None:
    ws = wb.create_sheet("HISTORY")
    ws.sheet_properties.tabColor = "808080"

    cols = [
        "SnapshotTimestamp", "BusinessDate", "ShiftName",
        "ShippedCartons", "TotalStaff", "ExpectedCartons",
        "PerformancePct", "RAG",
        "HRP_OpenCount", "Packed_OverdueCount", "AuditCount"
    ]
    header_style(ws, 1, cols)
    # Seed an empty row so the table range has at least 2 rows (openpyxl requirement)
    for c_idx in range(1, len(cols) + 1):
        ws.cell(row=2, column=c_idx, value="")
    add_table(ws, f"A1:{col_letter(len(cols))}2", "tblHistory", "TableStyleMedium2")
    _autofit(ws)


def build_data_quality(wb: Workbook) -> None:
    ws = wb.create_sheet("DATA_QUALITY")
    ws.sheet_properties.tabColor = "FFC000"

    ws["A1"] = "DATA QUALITY CHECKS"
    ws["A1"].font = Font(bold=True, size=14, color=_hex_font_color(HEADER_HEX))

    hdrs = ["Check", "Formula / Logic", "Result", "Status"]
    header_style(ws, 3, hdrs)

    checks = [
        ("Missing CartonID in HRP",
         '=COUNTIFS(tblHRP[CartonID],"")'),
        ("Future dates in IN_PACKED (LAST_PACKED)",
         '=COUNTIF(tblPacked[LAST_PACKED],">"&TODAY())'),
        ("Future dates in IN_SHIPPED_LPNS (BusinessDate)",
         '=COUNTIF(tblShipped[BusinessDate],">"&TODAY())'),
        ("Duplicate LPNs in Shipped list",
         "=COUNTIF(tblShipped[DupFlag],TRUE)"),
        ("StaffAvailable = 0 rows in Staffing",
         "=COUNTIF(tblStaffing[StaffAvailable],0)"),
        ('Invalid ShiftName in Shipped (not Day/Night)',
         '=SUMPRODUCT((tblShipped[ShiftName]<>"Day")*(tblShipped[ShiftName]<>"Night")*(tblShipped[ShiftName]<>""))'),
        ("Invalid ShiftName in Staffing",
         '=SUMPRODUCT((tblStaffing[ShiftName]<>"Day")*(tblStaffing[ShiftName]<>"Night")*(tblStaffing[ShiftName]<>""))'),
        ("HRP rows with IncludeInHRP=TRUE (open action items)",
         "=COUNTIF(tblHRP[IncludeInHRP],TRUE)"),
        ("Packed items where ActionFlag=TRUE (overdue)",
         "=COUNTIF(tblPacked[ActionFlag],TRUE)"),
    ]

    for r_off, (label, formula) in enumerate(checks, start=4):
        ws.cell(row=r_off, column=1, value=label)
        ws.cell(row=r_off, column=2, value=f"'{formula}")   # leading apostrophe forces Excel to treat this as display text, not an evaluated formula
        ws.cell(row=r_off, column=3).value = formula         # live formula
        ws.cell(row=r_off, column=4).value = \
            f'=IF(C{r_off}=0,"OK","WARNING")'

        for c in range(1, 5):
            ws.cell(row=r_off, column=c).border = thin_border()

        # CF: WARNING = amber
        ws.conditional_formatting.add(
            f"D{r_off}:D{r_off}",
            Rule(type="containsText", operator="containsText", text="WARNING",
                 dxf=DifferentialStyle(fill=fill(AMBER_HEX), font=Font(bold=True)))
        )
        ws.conditional_formatting.add(
            f"D{r_off}:D{r_off}",
            Rule(type="containsText", operator="containsText", text="OK",
                 dxf=DifferentialStyle(fill=fill(GREEN_HEX)))
        )

    _autofit(ws)


def build_charts(wb: Workbook) -> None:
    """Dedicated CHARTS sheet with 5 pre-built charts that update as data is added."""
    ws = wb.create_sheet("CHARTS")
    ws.sheet_properties.tabColor = "0070C0"

    ws["A1"] = "PERFORMANCE CHARTS  —  All charts update automatically as data is entered"
    ws["A1"].font = Font(bold=True, size=12, color=_hex_font_color(HEADER_HEX))

    ws_daily = wb["T_DISPATCH_DAILY"]
    ws_kpi   = wb["T_DISPATCH_KPI"]
    ws_hist  = wb["HISTORY"]

    # ── Chart 1: Daily Dispatch Performance % Trend (Line) ──────────────────
    chart1 = LineChart()
    chart1.title  = "Daily Dispatch Performance % Trend"
    chart1.style  = 10
    chart1.height = 14
    chart1.width  = 24
    chart1.y_axis.title  = "Performance %"
    chart1.x_axis.title  = "Business Date"
    chart1.y_axis.numFmt = "0%"
    chart1.y_axis.scaling.min = 0

    data1 = Reference(ws_daily, min_col=5, min_row=1, max_row=1048576)   # DailyPerformancePct
    chart1.add_data(data1, titles_from_data=True)
    cats1 = Reference(ws_daily, min_col=1, min_row=2, max_row=1048576)   # BusinessDate
    chart1.set_categories(cats1)

    s1 = chart1.series[0]
    s1.graphicalProperties.line.solidFill = BLUE_HEX[2:]
    s1.graphicalProperties.line.width = 20000
    s1.marker.symbol = "circle"
    s1.marker.size   = 5

    ws.add_chart(chart1, "A3")

    # ── Chart 2: Shipped vs Expected Cartons per Shift (Clustered Bar) ───────
    chart2 = BarChart()
    chart2.type     = "col"
    chart2.grouping = "clustered"
    chart2.title    = "Shipped vs Expected Cartons per Shift"
    chart2.style    = 10
    chart2.height   = 14
    chart2.width    = 24
    chart2.y_axis.title = "Cartons"
    chart2.x_axis.title = "Shift"

    data2a = Reference(ws_kpi, min_col=3, min_row=1, max_row=500)    # ShippedCartons
    data2b = Reference(ws_kpi, min_col=6, min_row=1, max_row=500)    # ExpectedCartons
    chart2.add_data(data2a, titles_from_data=True)
    chart2.add_data(data2b, titles_from_data=True)
    chart2.series[0].graphicalProperties.solidFill = BLUE_HEX[2:]
    chart2.series[1].graphicalProperties.solidFill = GREEN_HEX[2:]

    cats2 = Reference(ws_kpi, min_col=2, min_row=2, max_row=500)    # ShiftName
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "M3")

    # ── Chart 3: HRP Open & Packed Overdue Historical Trend (Bar) ───────────
    chart3 = BarChart()
    chart3.type     = "col"
    chart3.grouping = "clustered"
    chart3.title    = "HRP Open Items & Packed Overdue — Historical Trend"
    chart3.style    = 10
    chart3.height   = 14
    chart3.width    = 24
    chart3.y_axis.title = "Count"
    chart3.x_axis.title = "Snapshot Date"

    data3a = Reference(ws_hist, min_col=9,  min_row=1, max_row=500)  # HRP_OpenCount
    data3b = Reference(ws_hist, min_col=10, min_row=1, max_row=500)  # Packed_OverdueCount
    chart3.add_data(data3a, titles_from_data=True)
    chart3.add_data(data3b, titles_from_data=True)
    chart3.series[0].graphicalProperties.solidFill = RED_HEX[2:]
    chart3.series[1].graphicalProperties.solidFill = AMBER_HEX[2:]

    cats3 = Reference(ws_hist, min_col=2, min_row=2, max_row=500)   # BusinessDate
    chart3.set_categories(cats3)

    ws.add_chart(chart3, "A33")

    # ── Chart 4: Daily Staff Count Trend (Line) ───────────────────────────────
    chart4 = LineChart()
    chart4.title  = "Daily Staff Count Trend"
    chart4.style  = 10
    chart4.height = 14
    chart4.width  = 24
    chart4.y_axis.title = "Staff Count"
    chart4.x_axis.title = "Business Date"

    data4 = Reference(ws_daily, min_col=3, min_row=1, max_row=500)  # TotalStaff
    chart4.add_data(data4, titles_from_data=True)
    cats4 = Reference(ws_daily, min_col=1, min_row=2, max_row=500)
    chart4.set_categories(cats4)

    s4 = chart4.series[0]
    s4.graphicalProperties.line.solidFill = GREEN_HEX[2:]
    s4.graphicalProperties.line.width = 20000
    s4.marker.symbol = "square"
    s4.marker.size   = 5

    ws.add_chart(chart4, "M33")

    # ── Chart 5: Audit Performance % Trend (Line) ─────────────────────────────
    chart5 = LineChart()
    chart5.title  = "Audit Compliance % per Shift"
    chart5.style  = 10
    chart5.height = 14
    chart5.width  = 24
    chart5.y_axis.title  = "Audit %"
    chart5.x_axis.title  = "Shift"
    chart5.y_axis.numFmt = "0%"
    chart5.y_axis.scaling.min = 0

    data5 = Reference(ws_kpi, min_col=11, min_row=1, max_row=500)   # AuditPct
    chart5.add_data(data5, titles_from_data=True)
    cats5 = Reference(ws_kpi, min_col=2, min_row=2, max_row=500)
    chart5.set_categories(cats5)

    s5 = chart5.series[0]
    s5.graphicalProperties.line.solidFill = "70AD47"
    s5.graphicalProperties.line.width = 20000
    s5.marker.symbol = "diamond"
    s5.marker.size   = 5

    ws.add_chart(chart5, "A63")

    # Column A is narrow enough not to interfere with chart placement
    ws.column_dimensions["A"].width = 3
    _autofit(ws)


def build_dashboard(wb: Workbook) -> None:  # noqa: C901
    ws = wb.create_sheet("DASHBOARD")
    ws.sheet_properties.tabColor = "0070C0"

    # Light-grey page background (Excel-style)
    for r in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=14):
        for cell in r:
            cell.fill = fill(PAGE_BG_HEX)
            cell.font = Font(color=_hex_font_color(TEXT_DK_HEX))

    # ── Row 1: Title bar ──────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    tc = ws["A1"]
    tc.value = "WAREHOUSE / DISPATCH KPI DASHBOARD"
    tc.font = Font(bold=True, size=18, color=_hex_font_color(WHITE_HEX))
    tc.fill = fill(HEADER_HEX)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: Meta / last-refreshed bar ─────────────────────────────────────
    SUBTITLE_BG = "FFECF0F7"
    ws.merge_cells("A2:C2")
    _cell(ws, 2, 1, "Last Refreshed:", Font(bold=True, size=9,
          color=_hex_font_color(TEXT_MD_HEX)), fill(SUBTITLE_BG),
          Alignment(horizontal="right", vertical="center"))
    ws.merge_cells("D2:G2")
    _cell(ws, 2, 4, '=TEXT(NOW(),"dd/mm/yyyy hh:mm")',
          Font(bold=True, size=9, color=_hex_font_color(BLUE_HEX)),
          fill(SUBTITLE_BG), Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("H2:N2")
    _cell(ws, 2, 8, "Warehouse / Dispatch KPI System  v0.0.1-beta",
          Font(italic=True, size=9, color=_hex_font_color(TEXT_MD_HEX)),
          fill(SUBTITLE_BG), Alignment(horizontal="right", vertical="center"))
    ws.row_dimensions[2].height = 18

    # Row 3: thin visual spacer
    ws.row_dimensions[3].height = 6

    # ── Row 4: Section header "KPI OVERVIEW" ─────────────────────────────────
    _section_header(ws, 4, "  KPI OVERVIEW", SECN_HEX)
    ws.row_dimensions[4].height = 20

    # ── Rows 5-8: KPI cards — attention / alert row ───────────────────────────
    _kpi_card(ws, row=5, col=2, title="HRP OPEN ITEMS",
              formula="=COUNTIF(tblHRP[IncludeInHRP],TRUE)", accent=RED_HEX)
    _kpi_card(ws, row=5, col=6, title="PACKED OVERDUE",
              formula="=COUNTIF(tblPacked[ActionFlag],TRUE)", accent=RED_HEX)
    # Fixed formula: INDEX/MATCH returns #N/A (→ "N/A" via IFERROR) when no row
    # matches today's date in tblDispatchDaily.  The previous SUMIF returned 0,
    # which TEXT() formatted as "0.0%" — a misleading result when no data existed.
    _kpi_card(ws, row=5, col=10, title="DISPATCH PERF TODAY",
              formula='=IFERROR(TEXT(INDEX(tblDispatchDaily[DailyPerformancePct],MATCH(TODAY(),tblDispatchDaily[BusinessDate],0)),"0.0%"),"N/A")',
              accent=BLUE_HEX)

    # ── Rows 9-12: KPI cards — operations row ────────────────────────────────
    _kpi_card(ws, row=9, col=2, title="DUPLICATE LPNs",
              formula="=COUNTIF(tblShipped[DupFlag],TRUE)", accent=AMBER_HEX)
    _kpi_card(ws, row=9, col=6, title="STAFF TODAY",
              formula="=SUMIF(tblDispatchKPI[BusinessDate],TODAY(),tblDispatchKPI[TotalStaff])",
              accent=BLUE_HEX)
    _kpi_card(ws, row=9, col=10, title="CARTONS SHIPPED TODAY",
              formula="=SUMIF(tblDispatchKPI[BusinessDate],TODAY(),tblDispatchKPI[ShippedCartons])",
              accent=GREEN_HEX)

    # Row 13: spacer
    ws.row_dimensions[13].height = 8

    # ── Row 14: Section header "PERFORMANCE SUMMARY" ─────────────────────────
    _section_header(ws, 14, "  PERFORMANCE SUMMARY  (Most Recent Shifts)", SECN_HEX)
    ws.row_dimensions[14].height = 20

    # Row 15: table column headers
    perf_hdrs = ["Date", "Shift", "Shipped", "Expected", "Perf %", "RAG"]
    for ci, hdr in enumerate(perf_hdrs, start=1):
        c = ws.cell(row=15, column=ci, value=hdr)
        c.font = Font(bold=True, size=9, color=_hex_font_color(WHITE_HEX))
        c.fill = fill(BLUE_HEX)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[15].height = 16

    # Rows 16-18: 3 most-recent rows from tblDispatchKPI (newest first)
    #   ROWS(tblDispatchKPI[BusinessDate]) gives total row count;
    #   subtracting 0/1/2 offsets yields the last three rows.
    for row_off in range(3):
        rn = 16 + row_off
        idx_expr = f"ROWS(tblDispatchKPI[BusinessDate])-{row_off}"
        row_data = [
            f'=IFERROR(TEXT(INDEX(tblDispatchKPI[BusinessDate],{idx_expr}),"dd/mm/yyyy"),"")',
            f'=IFERROR(INDEX(tblDispatchKPI[ShiftName],{idx_expr}),"")',
            f'=IFERROR(INDEX(tblDispatchKPI[ShippedCartons],{idx_expr}),"")',
            f'=IFERROR(INDEX(tblDispatchKPI[ExpectedCartons],{idx_expr}),"")',
            f'=IFERROR(TEXT(INDEX(tblDispatchKPI[PerformancePct],{idx_expr}),"0.0%"),"")',
            f'=IFERROR(INDEX(tblDispatchKPI[RAG],{idx_expr}),"")',
        ]
        row_bg = CARD_BG_HEX if row_off % 2 == 0 else "FFF7F7F7"
        for ci, fml in enumerate(row_data, start=1):
            c = ws.cell(row=rn, column=ci, value=fml)
            c.font = Font(size=9, color=_hex_font_color(TEXT_DK_HEX))
            c.fill = fill(row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        # RAG conditional formatting for column F (col index 6)
        rag_ref = f"F{rn}"
        for text, hex_c, fnt in [
            ("Green", GREEN_HEX, Font(color=_hex_font_color(WHITE_HEX))),
            ("Amber", AMBER_HEX, Font(color=_hex_font_color(TEXT_DK_HEX))),
            ("Red",   RED_HEX,   Font(bold=True, color=_hex_font_color(WHITE_HEX))),
        ]:
            ws.conditional_formatting.add(
                rag_ref,
                Rule(type="containsText", operator="containsText", text=text,
                     dxf=DifferentialStyle(fill=fill(hex_c), font=fnt))
            )
        ws.row_dimensions[rn].height = 14

    # Row 19: spacer
    ws.row_dimensions[19].height = 8

    # ── Row 20: Section header "AUDIT OVERVIEW" ──────────────────────────────
    _section_header(ws, 20, "  AUDIT OVERVIEW  (Today)", "FF70AD47")
    ws.row_dimensions[20].height = 20

    # ── Rows 21-24: Audit KPI cards ──────────────────────────────────────────
    _kpi_card(ws, row=21, col=2, title="AUDITS TODAY",
              formula='=IFERROR(SUMIFS(tblAuditLog[AuditCount],tblAuditLog[BusinessDate],TODAY()),"N/A")',
              accent="FF70AD47")
    _kpi_card(ws, row=21, col=6, title="AUDIT TARGET (PER SHIFT)",
              formula='=IFERROR(INDEX(tblConfig_Rules[Value],MATCH("Audit_SampleSize",tblConfig_Rules[RuleName],0)),"N/A")',
              accent="FF70AD47")
    _kpi_card(ws, row=21, col=10, title="AUDIT COMPLIANCE TODAY",
              formula='=IFERROR(TEXT(SUMIFS(tblAuditLog[AuditCount],tblAuditLog[BusinessDate],TODAY())'
                      '/IFERROR(INDEX(tblConfig_Rules[Value],MATCH("Audit_SampleSize",tblConfig_Rules[RuleName],0)),1),"0.0%"),"N/A")',
              accent="FF70AD47")

    # Row 25: spacer
    ws.row_dimensions[25].height = 8

    # ── Row 26: Section header "RAG LEGEND" ───────────────────────────────────
    _section_header(ws, 26, "  RAG LEGEND", STEEL_HEX)
    ws.row_dimensions[26].height = 20

    # Row 27: legend colour blocks
    legend = [
        (2, "GREEN",  GREEN_HEX,  "Dispatch Performance >= 100%"),
        (6, "AMBER",  AMBER_HEX,  "Dispatch Performance >= 90% and < 100%"),
        (10, "RED",   RED_HEX,    "Dispatch Performance < 90%"),
    ]
    for start_c, label, hex_c, desc in legend:
        end_c = start_c + 2
        ws.merge_cells(f"{col_letter(start_c)}27:{col_letter(end_c)}27")
        lc = ws.cell(row=27, column=start_c, value=label)
        lc.font = Font(bold=True, size=9, color=_hex_font_color(WHITE_HEX))
        lc.fill = fill(hex_c)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = thin_border()

        ws.merge_cells(f"{col_letter(start_c)}28:{col_letter(end_c)}28")
        dc = ws.cell(row=28, column=start_c, value=desc)
        dc.font = Font(italic=True, size=8, color=_hex_font_color(TEXT_MD_HEX))
        dc.fill = fill(PAGE_BG_HEX)
        dc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[27].height = 16
    ws.row_dimensions[28].height = 14

    # Row 29: spacer
    ws.row_dimensions[29].height = 8

    # ── Row 30: Section header "QUICK NAVIGATION" ─────────────────────────────
    _section_header(ws, 30, "  QUICK NAVIGATION", NAV_HEX)
    ws.row_dimensions[30].height = 20

    # Row 31: navigation hyperlinks — Input / KPI sheets
    nav_links = [
        (2,  "IN_PACKED",         "#IN_PACKED!A1",         "004EA8"),
        (4,  "IN_HRP",            "#IN_HRP!A1",            "004EA8"),
        (6,  "IN_SHIPPED_LPNS",   "#IN_SHIPPED_LPNS!A1",   "006400"),
        (8,  "IN_STAFFING",       "#IN_STAFFING!A1",       "006400"),
        (10, "IN_AUDIT_LOG",      "#IN_AUDIT_LOG!A1",      "375623"),
        (12, "T_DISPATCH_KPI",    "#T_DISPATCH_KPI!A1",    "8B6914"),
        (14, "HISTORY",           "#HISTORY!A1",           "595959"),
    ]
    for start_ci, label, target, fc in nav_links:
        end_ci = start_ci + 1
        ws.merge_cells(f"{col_letter(start_ci)}31:{col_letter(end_ci)}31")
        nc = ws.cell(row=31, column=start_ci, value=f">> {label}")
        nc.hyperlink = target
        nc.font = Font(bold=True, size=8, underline="single", color=f"FF{fc}")
        nc.fill = fill(CARD_BG_HEX)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border = thin_border()
    ws.row_dimensions[31].height = 16

    # Row 32: secondary nav — analysis/chart sheets
    nav_links2 = [
        (2,  "CHARTS",           "#CHARTS!A1",           "0070C0"),
        (4,  "DATA_QUALITY",     "#DATA_QUALITY!A1",     "8B6914"),
        (6,  "ACTION_HRP",       "#ACTION_HRP!A1",       "8B0000"),
        (8,  "ACTION_PACKED",    "#ACTION_PACKED!A1",    "8B0000"),
    ]
    for start_ci, label, target, fc in nav_links2:
        end_ci = start_ci + 1
        ws.merge_cells(f"{col_letter(start_ci)}32:{col_letter(end_ci)}32")
        nc2 = ws.cell(row=32, column=start_ci, value=f">> {label}")
        nc2.hyperlink = target
        nc2.font = Font(bold=True, size=8, underline="single", color=f"FF{fc}")
        nc2.fill = fill(CARD_BG_HEX)
        nc2.alignment = Alignment(horizontal="center", vertical="center")
        nc2.border = thin_border()
    ws.row_dimensions[32].height = 16

    # Row 33: spacer
    ws.row_dimensions[33].height = 8

    # ── Row 34: Section header "CONTROLS" ────────────────────────────────────
    _section_header(ws, 34, "  CONTROLS  (VBA macro buttons — workbook must be saved as .xlsm with macros enabled)",
                    HEADER_HEX)
    ws.row_dimensions[34].height = 20

    # Row 35: macro-enable instructions
    ws.merge_cells("A35:N35")
    note = ws["A35"]
    note.value = ("⚠  To activate the buttons below: File → Save As → Excel Macro-Enabled Workbook (.xlsm). "
                  "Then enable macros when prompted on next open.  "
                  "Alternatively, run CreateKPIWorkbook.bas from the Developer → Visual Basic editor.")
    note.font = Font(italic=True, size=8, color=_hex_font_color(AMBER_HEX))
    note.fill = fill("FFFFF4CC")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[35].height = 28

    # Row 36: button descriptions
    btns = [
        (2,  "[ REFRESH ALL ]",            "Recalculates all formulas & refreshes PivotTables"),
        (6,  "[ TAKE DAILY SNAPSHOT ]",    "Appends current KPIs to the HISTORY sheet"),
        (10, "[ POPULATE ACTION SHEETS ]", "Copies filtered data to ACTION_HRP and ACTION_PACKED"),
    ]
    for ci, bname, bdesc in btns:
        ws.merge_cells(f"{col_letter(ci)}36:{col_letter(ci + 2)}36")
        bc = ws.cell(row=36, column=ci, value=bname)
        bc.font = Font(bold=True, size=9, color=_hex_font_color(HEADER_HEX))
        bc.fill = fill("FFFFF4CC")   # light yellow — mimics a button
        bc.alignment = Alignment(horizontal="center", vertical="center")
        bc.border = thin_border()

        ws.merge_cells(f"{col_letter(ci)}37:{col_letter(ci + 2)}37")
        dc = ws.cell(row=37, column=ci, value=bdesc)
        dc.font = Font(italic=True, size=8, color=_hex_font_color(TEXT_MD_HEX))
        dc.fill = fill(PAGE_BG_HEX)
        dc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[36].height = 18
    ws.row_dimensions[37].height = 24

    # ── Column widths ─────────────────────────────────────────────────────────
    # Col A = narrow left margin; B–N = even card-column width
    ws.column_dimensions["A"].width = 2
    for ci in range(2, 15):
        ws.column_dimensions[col_letter(ci)].width = 15


# ---------------------------------------------------------------------------
# Dashboard helper functions
# ---------------------------------------------------------------------------

def _cell(ws, row: int, col: int, value, font, cell_fill, alignment) -> None:
    """Write a value + style to a single cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.fill = cell_fill
    c.alignment = alignment


def _section_header(ws, row: int, text: str, bg: str) -> None:
    """Full-width (A:N) section header bar."""
    ws.merge_cells(f"A{row}:N{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(bold=True, size=11, color=_hex_font_color(WHITE_HEX))
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _kpi_card(ws, row: int, col: int, title: str, formula: str,
              accent: str) -> None:
    """Excel-style KPI card: coloured title bar (1 row) + white value area (3 rows).

    Layout (4 rows total starting at *row*):
      row   – coloured accent bar with white title label
      row+1 – ┐
      row+2 – ┤ merged value cell: large coloured number
      row+3 – ┘
    Each card is 3 columns wide.
    """
    CARD_W = 3
    end_col = col + CARD_W - 1
    thin = Side(style="thin", color=accent[2:])   # accent-coloured thin border

    # Title bar
    ws.merge_cells(f"{col_letter(col)}{row}:{col_letter(end_col)}{row}")
    tc = ws.cell(row=row, column=col, value=title)
    tc.font = Font(bold=True, size=9, color=_hex_font_color(WHITE_HEX))
    tc.fill = fill(accent)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18

    # Value area
    ws.merge_cells(f"{col_letter(col)}{row+1}:{col_letter(end_col)}{row+3}")
    vc = ws.cell(row=row + 1, column=col, value=formula)
    vc.font = Font(bold=True, size=24, color=_hex_font_color(accent))
    vc.fill = fill(CARD_BG_HEX)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row + 1].height = 42
    ws.row_dimensions[row + 2].height = 8
    ws.row_dimensions[row + 3].height = 8

    # Accent-coloured border around the whole card
    for r in range(row, row + 4):
        for c in range(col, end_col + 1):
            left_side  = thin if c == col     else None
            right_side = thin if c == end_col else None
            top_side   = thin if r == row     else None
            bot_side   = thin if r == row + 3 else None
            if any([left_side, right_side, top_side, bot_side]):
                ws.cell(row=r, column=c).border = Border(
                    left=left_side, right=right_side,
                    top=top_side, bottom=bot_side
                )


def _autofit(ws) -> None:
    """Approximate auto-fit (openpyxl does not have native auto-fit)."""
    for col in ws.columns:
        max_len = 0
        col_letter_str = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter_str].width = min(max(max_len + 2, 10), 40)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_workbook(output_path: str = "KPI_Workbook.xlsm") -> None:
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Build all sheets
    build_config(wb)
    build_in_packed(wb)
    build_in_hrp(wb)
    build_in_shipped(wb)
    build_in_staffing(wb)
    build_in_targets(wb)
    build_in_audit_log(wb)
    build_t_dispatch_kpi(wb)
    build_t_dispatch_daily(wb)
    build_action_hrp(wb)
    build_action_packed(wb)
    build_history(wb)
    build_data_quality(wb)
    build_charts(wb)
    build_dashboard(wb)

    # Re-order sheets for usability.
    # Iterates desired order; skips any sheet not present to avoid KeyError.
    desired_order = [
        "DASHBOARD", "CHARTS", "CONFIG",
        "IN_PACKED", "IN_HRP", "IN_SHIPPED_LPNS", "IN_STAFFING",
        "IN_TARGETS_DAILY", "IN_AUDIT_LOG",
        "T_DISPATCH_KPI", "T_DISPATCH_DAILY",
        "ACTION_HRP", "ACTION_PACKED",
        "HISTORY", "DATA_QUALITY"
    ]
    missing = [n for n in desired_order if n not in wb.sheetnames]
    if missing:
        print(f"Warning: the following sheets are missing and will be skipped in reordering: {missing}")
    for idx, name in enumerate(desired_order):
        if name not in wb.sheetnames:
            continue
        current_pos = wb.sheetnames.index(name)
        if current_pos != idx:
            wb.move_sheet(name, offset=idx - current_pos)

    wb.save(output_path)
    print(f"Workbook saved: {output_path}")
    print(f"Unprotect password: {UNPROTECT_PW}")
    print("\nSheets created:")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    build_workbook()
